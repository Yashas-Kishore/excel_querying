# core.py

import openpyxl
import json

def load_excel_file(file_path):
    """Load an Excel file."""
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading file: {e}")
        return None
    return wb

def convert_excel_to_json(file_path):
    """Convert Excel data to JSON format."""
    wb = load_excel_file(file_path)
    if wb is None:
        return None, None

    data = {}
    sheet_columns = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        worksheet_data = []

        headers = [header.strip() for header in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        sheet_columns[sheet_name] = headers

        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {}
            for i in range(len(headers)):
                row_data[headers[i]] = str(row[i]) if row[i] is not None else ''
            worksheet_data.append(row_data)

        data[sheet_name] = worksheet_data

    json_data = json.dumps(data, indent=4)

    with open('all_data.json', 'w') as f:
        f.write(json_data)

    print("Converted all worksheets to JSON")
    return sheet_columns, data

def display_attributes(sheet_columns):
    """Display the available attributes in the data."""
    print("\nAvailable attributes in the data:")
    for sheet_name, columns in sheet_columns.items():
        print(f"\nSheet '{sheet_name}' has the following columns:")
        for col in columns:
            print(f" - {col}")

def parse_query(query, sheet_columns):
    """Parse the user query to extract column and value."""
    query = query.lower()
    
    # Extract column name and value from the query
    column = None
    value = None
    
    # Look for column names in the query
    for sheet_name, columns in sheet_columns.items():
        for col in columns:
            if col.lower() in query:
                column = col
                break
    
    if column:
        # Extract the value from the query
        # Simplified approach to get the value from the query
        column_index = query.lower().find(column.lower()) + len(column)
        remaining_query = query[column_index:].strip()
        
        # Assume value is the first segment after the column name
        if remaining_query.startswith("as"):
            value = remaining_query[2:].strip().strip("'\"")
    
    return column, value

def query_data(sheet_columns, data):
    """Process user queries and return matching data."""
    print("\nWelcome to the JSON data query interface!")
    display_attributes(sheet_columns)  # Show available attributes (column names) to the user

    print("\nYou can ask queries about the data by typing anything.")
    print("For example: 'students with gender as male'. Type 'exit' to quit.")

    while True:
        query = input("\nEnter your query: ").lower()
        if query == 'exit':
            break
        
        column, value = parse_query(query, sheet_columns)
        
        if column and value:
            found = False
            seen_rows = set()
            for sheet_name, worksheet_data in data.items():
                for row in worksheet_data:
                    if column in row and value == row[column].lower():
                        row_tuple = tuple(row.items())
                        if row_tuple not in seen_rows:
                            seen_rows.add(row_tuple)
                            print(f"Found a match in worksheet '{sheet_name}' at column '{column}'!")
                            print(row)
                            found = True
            if not found:
                print("No matches found.")
        else:
            print("Unable to parse the query. Make sure the query contains a valid column and value.")