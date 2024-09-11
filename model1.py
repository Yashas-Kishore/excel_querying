import openpyxl
import json

def load_excel_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error loading file: {e}")
        return None
    return wb

def convert_excel_to_json(file_path):
    wb = load_excel_file(file_path)
    if wb is None:
        return

    # Create a dictionary to store the data
    data = {}

    # Store the attributes (column names) for each sheet
    sheet_columns = {}

    # Iterate over all worksheets in the file
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Create a list to store the data for this worksheet
        worksheet_data = []

        # Get the column headers (first row)
        headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

        # Store the column headers for the current sheet
        sheet_columns[sheet_name] = headers

        # Iterate over the remaining rows in the sheet (starting from the second row)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Create a dictionary to store the row data
            row_data = {}

            # Iterate over the columns in the row and map them to their respective header
            for i in range(len(headers)):
                row_data[headers[i]] = str(row[i]) if row[i] is not None else ''

            # Add the row data to the worksheet data
            worksheet_data.append(row_data)

        # Add the worksheet data to the main dictionary
        data[sheet_name] = worksheet_data

    # Convert the data to JSON
    json_data = json.dumps(data, indent=4)

    # Save the JSON data to a file
    with open('all_data.json', 'w') as f:
        f.write(json_data)

    print("Converted all worksheets to JSON")
    return sheet_columns, data

def display_attributes(sheet_columns):
    print("\nAvailable attributes in the data:")
    for sheet_name, columns in sheet_columns.items():
        print(f"\nSheet '{sheet_name}' has the following columns:")
        for col in columns:
            print(f" - {col}")

def query_data(sheet_columns, data):
    print("\nWelcome to the JSON data query interface!")
    display_attributes(sheet_columns)  # Show available attributes (column names) to the user

    print("\nYou can specify column names and values to search.")
    print("For example, first select a column, then provide the value you want to search for in that column.")
    print("You can search multiple columns and values at once. Type 'exit' to quit.")

    while True:
        search_criteria = {}

        # Loop to get user input for multiple column-value pairs
        while True:
            column = input("\nEnter a column name to search (or type 'done' to finish): ").strip()
            if column.lower() == 'done':
                break
            elif column.lower() == 'exit':
                return
            else:
                # Check if the column exists in any of the sheets
                available_columns = set(col for cols in sheet_columns.values() for col in cols)
                if column not in available_columns:
                    print(f"Column '{column}' not found. Please select from available columns.")
                    display_attributes(sheet_columns)
                    continue

                value = input(f"Enter the value to search for in column '{column}': ").strip().lower()
                search_criteria[column] = value

        if not search_criteria:
            print("No search criteria provided. Please enter at least one column-value pair.")
            continue

        found = False
        seen_rows = set()  # To keep track of already printed rows
        for worksheet_name, worksheet_data in data.items():
            for row in worksheet_data:
                # Check if all column-value pairs match for the row
                if all(search_criteria[col].lower() in (str(row[col]).lower() if row[col] else '') for col in search_criteria):
                    row_tuple = tuple(row.items())  # Convert row to a tuple to be hashable
                    if row_tuple not in seen_rows:
                        seen_rows.add(row_tuple)
                        print(f"\nFound a match in worksheet '{worksheet_name}'!")
                        print(row)
                        found = True

        if not found:
            print("No matches found.")

def main():
    file_path = input("Enter the path to the Excel file: ")
    sheet_columns, data = convert_excel_to_json(file_path)
    if sheet_columns and data:
        query_data(sheet_columns, data)

if __name__ == "__main__":
    main()
